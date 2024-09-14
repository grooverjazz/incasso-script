# Zorgt ervoor dat we met CSV kunnen werken.
import csv
import openpyxl
import openpyxl.worksheet
import os.path
import copy
from pycel import ExcelCompiler
from decimal import *

# Filter bullshit
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


# Declareer types (scheelt in leesbaarheid)
IncassoSheet = dict[str, dict[str, any]]
IncassoFields = list[str]


# De naam van het lidnummer- en naamveld.
FIELD_ID = "id"
FIELD_NAME = "naam"


# (Importeert een CSV-bestand)
def import_csv(file_name: str) -> tuple[IncassoSheet, IncassoFields]:
    print(f"{file_name} openen...")
    
    with open(file_name, 'r', encoding='utf-8-sig', newline='') as input_file:
        file_reader: csv.DictReader = csv.DictReader(input_file, delimiter=";")

        return ({row[FIELD_ID]: row for row in file_reader}, file_reader.fieldnames)


# (Importeert een Excel-bestand)
def import_excel(file_name: str) -> tuple[IncassoSheet, IncassoFields]:
    # Open de sheet, maak een compiler aan
    workbook: openpyxl.Workbook = openpyxl.load_workbook(file_name)
    excel = ExcelCompiler(excel=workbook)

    # Pak de goeie sheet-naam
    sheet_name: str = "Incasso" if "Incasso" in workbook.sheetnames else workbook.sheetnames[0]
    sheet: openpyxl.worksheet = workbook[sheet_name]

    # Verkrijg de fields
    rows: list[any] = list(sheet.rows)
    fields: IncassoFields = [str(v.value) for v in rows.pop(0) if v != None]
    
    # Hack: soms doet Excel kut
    fields = [field for field in fields if field != "None"]

    # (Verwerkt een Excel-cel)
    def process(val: any, fieldname: str) -> any:
        # Evalueer formule indien nodig
        if str(val).startswith('='): val = excel.evaluate(str(val)[1:])
        
        # Zet ID's naar string
        if fieldname == FIELD_ID: val = str(int(val))

        return val

    # Ga door alle rijen
    res: IncassoSheet = {}
    for row in rows:
        # Verwerk rij naar een dict
        row_dict: dict = {
            fieldname: process(cell.value, fieldname)
                for (cell, fieldname) in zip(list(row), fields)
        }

        # Verkrijg id, zeik als 'ie al bestaat
        id: str = row_dict[FIELD_ID]
        assert id not in res, f"Lid {id} staat dubbel in de incasso!"

        # Schrijf naar resultaat
        res[id] = row_dict
    
    # Return resultaat en fields
    return (res, fields)


# (Vertaalt een prijs van tekst naar een getal)
def str_to_cents(price_str: str) -> int:
    _price_str = price_str

    # Geen euros! Geen komma's! Geen spaties!
    price_str = price_str\
        .replace('€', '')\
        .replace(',', '.')\
        .replace(' ', '')
    
    # Een leeg bedrag is 0 euro
    if price_str == "": return 0

    # Probeer te parsen naar een Decimal
    #  (vermenigvuldig met 100, maak er een int van)
    try:
        return int(Decimal(price_str).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) * 100)
    except:
        raise Exception(f"Kan prijs '{_price_str}' niet omzetten naar een getal!")


# (Vertaalt een prijs van getal naar tekst)
def cents_to_str(price: int) -> str:
    total_euro = price // 100
    total_cent = price % 100

    return f"{total_euro},{total_cent:02d}"


# (Voegt meerdere lijsten samen in de ledenadmin)
def add_lists_to_ledenadmin(
    ledenadmin: tuple[IncassoSheet, IncassoFields],
    lists: list[tuple[IncassoSheet, IncassoFields]],
    totaal: str,
    beschrijving: str,
    depth: int = 0,
) -> None:
    (ledenadmin_members, ledenadmin_fields) = ledenadmin

    # Maak een leeg totaal en beschrijving aan in de ledenadmin
    for member_id in ledenadmin_members:
        member = ledenadmin_members[member_id]
        member[totaal] = 0
        member[beschrijving] = ""
    ledenadmin_fields += [totaal, beschrijving]

    # Ga door alle lijsten
    for (list_members, list_fields) in lists:
        # Verkrijg alle nieuwe fields
        new_fields: IncassoFields = []
        for field in list_fields:
            if field.strip() == "": continue
            if field in ledenadmin_fields: continue

            new_fields.append(field)
        
        # Maak de naam van de lijst (voor debug)
        list_name: str = ", ".join(new_fields)
        print(depth * "   " + f"{list_name} samenvoegen...")

        # Ga door alle leden in de lijst
        for member_id in list_members:
            # Pak de lid-info uit de lijst
            list_member: str[dict, any] = list_members[member_id]

            # Check of het lid in de ledenadmin staat
            if not (member_id in ledenadmin_members):
                print((depth + 1) * "   " + f"Lid {list_member[FIELD_NAME]} ({member_id}) staat niet in de ledenadmin!")
                member_id = "-1"

            # Pak de lid-info uit de ledenadmin
            ledenadmin_member: dict[str, any] = ledenadmin_members[member_id]
            
            # Ga door alle kolommen van de lijst
            for fieldname in new_fields:
                # Verkrijg een prijs en sla het incassoveld misschien over
                price: int = str_to_cents(str(list_member[fieldname]))
                if price == 0: continue

                # Voeg de prijs toe aan het lid
                ledenadmin_member[totaal] += price

                # Verkrijg een beschrijving
                price_str_legible: str = "€" + cents_to_str(price)
                description: str = f"{price_str_legible} ({field})"

                # Vul de beschrijving aan als het een 'onbekend' lid is
                if member_id == "-1":
                    description += f"[{list_member[FIELD_ID]}; {list_member[FIELD_NAME]}]"

                # Voeg de beschrijving toe aan het lid
                #   (en een kommaatje waar nodig)
                if ledenadmin_member[beschrijving] != "":
                    ledenadmin_member[beschrijving] += ", "
                
                ledenadmin_member[beschrijving] += description
    
    # Maak strings van totalen
    for member_id in ledenadmin_members:
        member: dict[str, any] = ledenadmin_members[member_id]
        member[totaal] = cents_to_str(member[totaal])


# Definieer de ledenadmin die gebruikt gaat worden voor samenvoegen
ledenadmin_members: IncassoSheet
ledenadmin_fields: IncassoFields

# (Zet de ledenadmin vanuit de notebook)
def set_ledenadmin(new_ledenadmin_members: IncassoSheet, new_ledenadmin_fields: IncassoFields) -> None:
    global ledenadmin_members, ledenadmin_fields
    ledenadmin_members = new_ledenadmin_members
    ledenadmin_fields = new_ledenadmin_fields


# (Maakt een lijst van alle bestanden in een map)
def merge_directory(directory: str, totaal: str, beschrijving: str, depth: int = 0) -> tuple[IncassoSheet, IncassoFields]:
    print(depth * "   " + f"{totaal} verwerken...")

    # Maak eerst lijsten van alle inhoud van de map
    #  (opmerking: zie hier de (recursieve) call naar `merge_file`!)
    dir_lists: list[tuple[IncassoSheet, IncassoFields]] = list(
        l for l in [merge_file(directory + "/" + file, depth + 1) for file in os.listdir(directory)] if l != None
    )

    # Maak een kopie van de ledenadmin (voor de zekerheid I guess)
    ledenadmin_copy: IncassoSheet = copy.deepcopy(ledenadmin_members)
    fields_copy: IncassoFields = copy.deepcopy(ledenadmin_fields)

    # Voeg alle inhoud van de map toe aan de ledenadmin
    add_lists_to_ledenadmin((ledenadmin_copy, fields_copy), dir_lists, totaal, beschrijving, depth + 1)

    return (ledenadmin_copy, fields_copy)


# (Maakt een lijst van een bestand (of een map))
def merge_file(file: str, depth: int = 0) -> tuple[IncassoSheet, IncassoFields] | None:
    print(depth * "   " + f"{file} openen...")

    if file.endswith(".xlsx"):
        # Importeer het Excel-bestand
        return import_excel(file)
    
    elif os.path.isdir(file):
        # Verkrijg de naam van de map vanuit z'n pad
        basename: str = os.path.basename(file)

        # Voeg alle bestanden samen vanuit de map
        merged: tuple[IncassoSheet, IncassoFields] = merge_directory(file, basename, '__temp_beschrijving', depth + 1)
        (merged_members, merged_fields) = merged

        # Verwijder de tijdelijke beschrijving
        merged_fields.remove("__temp_beschrijving")
        for member_id in merged_members:
            member = merged_members[member_id]
            member.pop("__temp_beschrijving")
        
        return merged

    else:
        # Print foutmelding
        print(f"Bestand {file} overgeslagen!")
        return None